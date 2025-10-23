#!/usr/bin/env python3
"""Excel 형태의 샘플 ORAN 메트릭 데이터 생성 스크립트.

CFM 담당자에게 전달할 샘플 데이터를 Excel 형식으로 생성합니다.
세 개의 Sheet로 구성:
- Sheet 1: 메트릭 데이터 (실제 샘플)
- Sheet 2: 필드 설명
- Sheet 3: 예상 이상 케이스
"""

import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
import sys

# 프로젝트 루트를 경로에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))


def create_metric_data_sheet():
    """Sheet 1: 메트릭 데이터 생성."""

    # CSV 파일에서 데이터 읽기
    csv_path = Path(__file__).parent.parent / "data" / "samples" / "sample_oran_metrics_wide.csv"

    if csv_path.exists():
        df = pd.read_csv(csv_path)
        return df

    # CSV가 없으면 직접 생성
    data = []
    base_time = datetime(2025, 10, 22, 9, 0, 0)

    # o-ru-001: Drift 패턴 (정상 → 점진적 증가 → Critical → 복구)
    scenarios = [
        # 정상 운영
        (0, "o-ru-001", "Tower-A", "Urban", 5.2, 102.3, 7.1, True, 1000, 0, "정상 운영 중"),
        (1, "o-ru-001", "Tower-A", "Urban", 5.4, 105.2, 7.3, True, 1000, 0, "정상"),
        (2, "o-ru-001", "Tower-A", "Urban", 5.1, 100.8, 7.0, True, 1000, 0, "정상"),
        (3, "o-ru-001", "Tower-A", "Urban", 5.3, 103.5, 7.2, True, 1000, 0, "정상"),
        (4, "o-ru-001", "Tower-A", "Urban", 5.5, 106.1, 7.4, True, 1000, 0, "정상"),
        # Drift 시작
        (5, "o-ru-001", "Tower-A", "Urban", 8.2, 158.3, 10.5, True, 1000, 0, "⚠️ RTT 증가 시작"),
        (6, "o-ru-001", "Tower-A", "Urban", 12.5, 215.7, 14.8, True, 1000, 0, "⚠️ RTT 계속 증가 (Drift)"),
        (7, "o-ru-001", "Tower-A", "Urban", 15.3, 255.9, 18.2, False, 1000, 0, "⚠️ RTT 높음 + LBM 실패"),
        # Critical
        (8, "o-ru-001", "Tower-A", "Urban", 25.8, 350.1, 25.5, False, 1000, 1, "🚨 CRITICAL: 높은 지연 + LBM 실패 + CCM 누락"),
        (9, "o-ru-001", "Tower-A", "Urban", 26.2, 355.3, 26.1, False, 1000, 2, "🚨 CRITICAL: 이상 지속 중"),
        # 복구
        (10, "o-ru-001", "Tower-A", "Urban", 18.5, 280.5, 19.3, False, 1000, 1, "⚠️ 회복 중"),
        (11, "o-ru-001", "Tower-A", "Urban", 10.2, 180.2, 12.1, True, 1000, 0, "⚠️ 회복 중"),
        (12, "o-ru-001", "Tower-A", "Urban", 6.8, 125.3, 8.5, True, 1000, 0, "거의 정상"),
        (13, "o-ru-001", "Tower-A", "Urban", 5.3, 103.1, 7.2, True, 1000, 0, "✅ 정상 복구"),
        (14, "o-ru-001", "Tower-A", "Urban", 5.1, 101.2, 7.0, True, 1000, 0, "정상"),

        # o-ru-002: Spike 패턴 (일시적 급증)
        (0, "o-ru-002", "Tower-B", "Rural", 4.8, 98.5, 6.9, True, 1000, 0, "정상 운영 중"),
        (1, "o-ru-002", "Tower-B", "Rural", 4.9, 99.2, 7.0, True, 1000, 0, "정상"),
        (2, "o-ru-002", "Tower-B", "Rural", 4.7, 97.8, 6.8, True, 1000, 0, "정상"),
        (3, "o-ru-002", "Tower-B", "Rural", 4.8, 98.3, 6.9, True, 1000, 0, "정상"),
        (4, "o-ru-002", "Tower-B", "Rural", 22.5, 320.5, 21.8, False, 1000, 0, "🚨 갑작스런 Spike 발생"),
        (5, "o-ru-002", "Tower-B", "Rural", 5.0, 100.1, 7.1, True, 1000, 0, "✅ 즉시 정상 복구 (일시적 spike)"),
        (6, "o-ru-002", "Tower-B", "Rural", 4.8, 98.7, 6.9, True, 1000, 0, "정상"),

        # o-ru-003: Jitter 패턴 (불규칙한 변동)
        (0, "o-ru-003", "Tower-C", "Suburban", 5.5, 108.2, 7.5, True, 1000, 0, "정상 운영 중"),
        (1, "o-ru-003", "Tower-C", "Suburban", 5.3, 105.1, 7.3, True, 1000, 0, "정상"),
        (2, "o-ru-003", "Tower-C", "Suburban", 5.6, 109.3, 7.6, True, 1000, 0, "정상"),
        (3, "o-ru-003", "Tower-C", "Suburban", 8.2, 155.2, 10.1, True, 1000, 0, "약간 증가"),
        (4, "o-ru-003", "Tower-C", "Suburban", 15.3, 250.8, 18.2, True, 1000, 0, "⚠️ Jitter 발생 (불안정)"),
        (5, "o-ru-003", "Tower-C", "Suburban", 6.1, 112.5, 7.8, True, 1000, 0, "다시 낮아짐"),
        (6, "o-ru-003", "Tower-C", "Suburban", 18.8, 285.3, 22.1, True, 1000, 0, "⚠️ 다시 높아짐 (Jitter)"),
        (7, "o-ru-003", "Tower-C", "Suburban", 5.4, 106.2, 7.4, True, 1000, 0, "다시 정상"),
        (8, "o-ru-003", "Tower-C", "Suburban", 5.5, 108.1, 7.5, True, 1000, 0, "정상"),

        # o-du-001: DU 정상 동작 (낮은 지연)
        (0, "o-du-001", "Datacenter-A", "Urban", 3.2, 85.2, 4.5, True, 1000, 0, "정상 운영 중 (DU는 지연 낮음)"),
        (1, "o-du-001", "Datacenter-A", "Urban", 3.1, 84.8, 4.4, True, 1000, 0, "정상"),
        (2, "o-du-001", "Datacenter-A", "Urban", 3.3, 86.1, 4.6, True, 1000, 0, "정상"),
        (3, "o-du-001", "Datacenter-A", "Urban", 3.2, 85.5, 4.5, True, 1000, 0, "정상"),
    ]

    for offset, endpoint_id, site_name, zone, udp_rtt, ecpri, lbm_rtt, lbm_success, ccm_int, ccm_miss, notes in scenarios:
        timestamp = base_time + timedelta(seconds=offset)
        data.append({
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "endpoint_id": endpoint_id,
            "site_name": site_name,
            "zone": zone,
            "udp_echo_rtt_ms": udp_rtt,
            "ecpri_delay_us": ecpri,
            "lbm_rtt_ms": lbm_rtt,
            "lbm_success": lbm_success,
            "ccm_interval_ms": ccm_int,
            "ccm_miss_count": ccm_miss,
            "notes": notes
        })

    return pd.DataFrame(data)


def create_field_description_sheet():
    """Sheet 2: 필드 설명 생성."""

    descriptions = [
        {
            "필드명": "timestamp",
            "레벨": "필수 (Level 1)",
            "데이터 타입": "문자열 (YYYY-MM-DD HH:MM:SS)",
            "설명": "메트릭 측정 시각. 시스템 전체에서 통일된 시간 기준 필요 (UTC 또는 로컬 시간 명시)",
            "예시 값": "2025-10-22 09:00:00",
            "비고": "시간 동기화가 중요함 (NTP 사용 권장)"
        },
        {
            "필드명": "endpoint_id",
            "레벨": "필수 (Level 1)",
            "데이터 타입": "문자열",
            "설명": "장비의 고유 식별자. O-RU 또는 O-DU를 구분할 수 있어야 함",
            "예시 값": "o-ru-001, o-du-002",
            "비고": "명명 규칙 정의 필요"
        },
        {
            "필드명": "site_name",
            "레벨": "선택 (Level 3)",
            "데이터 타입": "문자열",
            "설명": "설치 위치 이름 (사람이 읽을 수 있는 형태)",
            "예시 값": "Tower-A, Datacenter-B",
            "비고": "지역별 분석에 유용"
        },
        {
            "필드명": "zone",
            "레벨": "선택 (Level 3)",
            "데이터 타입": "문자열",
            "설명": "설치 지역 타입 (도심/교외/농촌 등)",
            "예시 값": "Urban, Suburban, Rural",
            "비고": "지역 특성별 임계값 조정 가능"
        },
        {
            "필드명": "udp_echo_rtt_ms",
            "레벨": "필수 (Level 1)",
            "데이터 타입": "실수",
            "설명": "UDP Echo 왕복 시간 (Round Trip Time). 기본적인 네트워크 지연 측정",
            "예시 값": "5.2 (정상), 25.8 (이상)",
            "비고": "정상: 3-7ms, 경고: 10-20ms, 위험: 20ms 이상"
        },
        {
            "필드명": "ecpri_delay_us",
            "레벨": "권장 (Level 2)",
            "데이터 타입": "실수",
            "설명": "eCPRI 프로토콜 지연 시간 (마이크로초). ORAN Fronthaul 지연 측정",
            "예시 값": "102.3 (정상), 350.1 (이상)",
            "비고": "정상: 80-120us, 경고: 200-300us, 위험: 300us 이상"
        },
        {
            "필드명": "lbm_rtt_ms",
            "레벨": "권장 (Level 2)",
            "데이터 타입": "실수",
            "설명": "CFM Loopback Message 왕복 시간. Ethernet Layer 연결성 검증",
            "예시 값": "7.1 (정상), 25.5 (이상)",
            "비고": "정상: 5-10ms, 경고: 10-20ms, 위험: 20ms 이상"
        },
        {
            "필드명": "lbm_success",
            "레벨": "권장 (Level 2)",
            "데이터 타입": "불리언 (TRUE/FALSE)",
            "설명": "Loopback Message 성공 여부. 연결성 확인",
            "예시 값": "TRUE (정상), FALSE (실패)",
            "비고": "FALSE는 심각한 문제 지표"
        },
        {
            "필드명": "ccm_interval_ms",
            "레벨": "권장 (Level 2)",
            "데이터 타입": "정수",
            "설명": "Continuity Check Message 전송 간격",
            "예시 값": "1000 (1초마다)",
            "비고": "일반적으로 고정값"
        },
        {
            "필드명": "ccm_miss_count",
            "레벨": "권장 (Level 2)",
            "데이터 타입": "정수",
            "설명": "CCM 누락 횟수. 연속성 확인 실패",
            "예시 값": "0 (정상), 1-2 (경고), 3+ (위험)",
            "비고": "3회 이상 누락 시 연결 끊김 의심"
        },
        {
            "필드명": "notes",
            "레벨": "선택 (Level 3)",
            "데이터 타입": "문자열",
            "설명": "사람이 읽을 수 있는 메모 (선택사항)",
            "예시 값": "정상, ⚠️ 경고, 🚨 위험",
            "비고": "디버깅 및 분석에 유용"
        },
    ]

    return pd.DataFrame(descriptions)


def create_anomaly_cases_sheet():
    """Sheet 3: 예상 이상 케이스 생성."""

    cases = [
        {
            "케이스 번호": "1",
            "이상 유형": "Drift (점진적 증가)",
            "설명": "메트릭이 서서히 증가하여 정상 범위를 벗어남. 장비 성능 저하 또는 부하 증가 시 발생",
            "예상 원인": "- 네트워크 부하 증가\n- 장비 온도 상승\n- 메모리 부족\n- 대역폭 포화",
            "탐지 방법": "CUSUM (누적합) 또는 PELT (변화점 탐지) 알고리즘",
            "대응 방안": "- 부하 재분배\n- 장비 점검\n- 용량 증설 검토",
            "예시 데이터": "o-ru-001의 09:00:05 ~ 09:00:09 구간\nUDP RTT: 5.2 → 8.2 → 12.5 → 25.8 ms"
        },
        {
            "케이스 번호": "2",
            "이상 유형": "Spike (급격한 증가)",
            "설명": "메트릭이 갑자기 급증했다가 즉시 정상으로 복귀. 일시적인 네트워크 혼잡",
            "예상 원인": "- 일시적인 네트워크 혼잡\n- 버퍼 오버플로우\n- 순간적인 간섭\n- 패킷 재전송",
            "탐지 방법": "통계적 이상값 탐지 (Z-score, IQR) 또는 임계값 기반",
            "대응 방안": "- 반복 발생 시 원인 조사\n- 단발성이면 모니터링만\n- QoS 정책 검토",
            "예시 데이터": "o-ru-002의 09:00:04 구간\nUDP RTT: 4.8 → 22.5 → 5.0 ms (즉시 복구)"
        },
        {
            "케이스 번호": "3",
            "이상 유형": "Jitter (불규칙한 변동)",
            "설명": "메트릭이 불규칙하게 오르락내리락. 네트워크 불안정 지표",
            "예상 원인": "- 경로 불안정 (라우팅 변경)\n- 간헐적인 간섭\n- 불안정한 전원\n- 동기화 문제",
            "탐지 방법": "표준편차 또는 변동 계수(CV) 분석, TCN/LSTM 예측 오차",
            "대응 방안": "- 네트워크 경로 안정화\n- 전원 상태 확인\n- 동기화 점검 (PTP/GPS)",
            "예시 데이터": "o-ru-003의 09:00:03 ~ 09:00:06 구간\nUDP RTT: 5.6 → 8.2 → 15.3 → 6.1 → 18.8 ms"
        },
        {
            "케이스 번호": "4",
            "이상 유형": "복합 장애 (Multi-metric)",
            "설명": "여러 메트릭이 동시에 이상. 심각한 장애 가능성",
            "예상 원인": "- 링크 단절\n- 장비 고장\n- 광 케이블 손상\n- 전원 장애",
            "탐지 방법": "다변량 탐지 (Isolation Forest, Multivariate Gaussian)",
            "대응 방안": "- 즉시 현장 점검\n- 백업 경로 활성화\n- 긴급 복구 절차 시작",
            "예시 데이터": "o-ru-001의 09:00:08 구간\nUDP RTT: 25.8ms (높음)\nLBM Success: FALSE (실패)\nCCM Miss: 1 (누락)"
        },
        {
            "케이스 번호": "5",
            "이상 유형": "정상 복구",
            "설명": "이상 상태에서 정상 상태로 복귀. 자동 복구 또는 수동 조치 후",
            "예상 원인": "- 일시적 문제 해결\n- 자동 재시작\n- 수동 조치 완료\n- 부하 감소",
            "탐지 방법": "이상 종료 탐지 (정상 범위 복귀 확인)",
            "대응 방안": "- 근본 원인 분석\n- 재발 방지 대책 수립\n- 모니터링 강화",
            "예시 데이터": "o-ru-001의 09:00:10 ~ 09:00:13 구간\nUDP RTT: 25.8 → 18.5 → 10.2 → 5.3 ms"
        },
    ]

    return pd.DataFrame(cases)


def main():
    """메인 함수."""
    print("\n" + "=" * 60)
    print("Excel 샘플 데이터 생성")
    print("=" * 60)

    # 출력 경로
    output_dir = Path(__file__).parent.parent / "data" / "samples"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / "sample_oran_metrics.xlsx"

    print(f"\n출력 파일: {output_file}")

    # Excel Writer 생성
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: 메트릭 데이터
        print("\n[1/3] Sheet 1: 메트릭 데이터 생성 중...")
        df_metrics = create_metric_data_sheet()
        df_metrics.to_excel(writer, sheet_name='메트릭 데이터', index=False)
        print(f"  - {len(df_metrics)}개 레코드 작성 완료")

        # Sheet 2: 필드 설명
        print("\n[2/3] Sheet 2: 필드 설명 생성 중...")
        df_fields = create_field_description_sheet()
        df_fields.to_excel(writer, sheet_name='필드 설명', index=False)
        print(f"  - {len(df_fields)}개 필드 설명 작성 완료")

        # Sheet 3: 예상 이상 케이스
        print("\n[3/3] Sheet 3: 예상 이상 케이스 생성 중...")
        df_cases = create_anomaly_cases_sheet()
        df_cases.to_excel(writer, sheet_name='예상 이상 케이스', index=False)
        print(f"  - {len(df_cases)}개 이상 케이스 작성 완료")

    # Excel 파일 스타일 조정
    print("\n[4/4] Excel 서식 조정 중...")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = load_workbook(output_file)

    # Sheet 1: 메트릭 데이터 서식
    ws1 = wb['메트릭 데이터']

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 열 너비 자동 조정
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: 필드 설명 서식
    ws2 = wb['필드 설명']

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 레벨에 따라 색상 구분
    level_colors = {
        "필수 (Level 1)": "FFE699",  # 노란색
        "권장 (Level 2)": "C5E0B4",  # 연두색
        "선택 (Level 3)": "D9E1F2",  # 하늘색
    }

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        level = row[1].value  # "레벨" 열
        if level in level_colors:
            fill = PatternFill(start_color=level_colors[level],
                             end_color=level_colors[level],
                             fill_type="solid")
            for cell in row:
                cell.fill = fill

        # 텍스트 줄바꿈
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 열 너비 조정
    ws2.column_dimensions['A'].width = 20  # 필드명
    ws2.column_dimensions['B'].width = 18  # 레벨
    ws2.column_dimensions['C'].width = 25  # 데이터 타입
    ws2.column_dimensions['D'].width = 60  # 설명
    ws2.column_dimensions['E'].width = 30  # 예시 값
    ws2.column_dimensions['F'].width = 35  # 비고

    # Sheet 3: 예상 이상 케이스 서식
    ws3 = wb['예상 이상 케이스']

    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 이상 유형별 색상
    case_colors = {
        "Drift (점진적 증가)": "F4B084",
        "Spike (급격한 증가)": "F8CBAD",
        "Jitter (불규칙한 변동)": "FFE699",
        "복합 장애 (Multi-metric)": "FF6B6B",
        "정상 복구": "C6EFCE",
    }

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        case_type = row[1].value  # "이상 유형" 열
        if case_type in case_colors:
            fill = PatternFill(start_color=case_colors[case_type],
                             end_color=case_colors[case_type],
                             fill_type="solid")
            row[1].fill = fill  # 이상 유형 셀만 색칠

        # 텍스트 줄바꿈
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 열 너비 조정
    ws3.column_dimensions['A'].width = 12  # 케이스 번호
    ws3.column_dimensions['B'].width = 25  # 이상 유형
    ws3.column_dimensions['C'].width = 50  # 설명
    ws3.column_dimensions['D'].width = 40  # 예상 원인
    ws3.column_dimensions['E'].width = 35  # 탐지 방법
    ws3.column_dimensions['F'].width = 40  # 대응 방안
    ws3.column_dimensions['G'].width = 45  # 예시 데이터

    # 행 높이 조정 (내용에 맞게)
    for row in ws3.iter_rows(min_row=2):
        ws3.row_dimensions[row[0].row].height = 80

    # 저장
    wb.save(output_file)

    print("\n" + "=" * 60)
    print("✅ Excel 파일 생성 완료!")
    print("=" * 60)
    print(f"\n📁 파일 위치: {output_file}")
    print(f"📊 총 3개 Sheet:")
    print(f"   1. 메트릭 데이터: {len(df_metrics)}개 레코드")
    print(f"   2. 필드 설명: {len(df_fields)}개 필드")
    print(f"   3. 예상 이상 케이스: {len(df_cases)}개 케이스")
    print(f"\n💡 이 파일을 CFM 담당자에게 전달하여 데이터 수집 가능 여부를 확인하세요.")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
